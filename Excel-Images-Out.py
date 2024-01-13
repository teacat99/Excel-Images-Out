from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from openpyxl.utils import get_column_letter, column_index_from_string
import os

#必要库安装
#pip install openpyxl
#pip install Pillow

def clear_screen():
    """清空屏幕"""
    # Windows
    if os.name == 'nt':
        os.system('cls')
    # Unix/Linux
    else:
        os.system('clear')

def create_unique_filename(base_path, extension):
    """生成一个唯一的文件名，如果文件名已存在，则添加计数器"""
    counter = 1
    new_path = base_path
    while os.path.exists(new_path + extension):
        new_path = f"{base_path}_{counter}"
        counter += 1
    return new_path + extension

def clean_path(input_path):
    """
    清洗路径字符串，处理多种情况下的特殊字符。
    """
    # 首先，特别处理 VSCode 拖拽文件路径时的格式（例如：& '路径'）
    if input_path.startswith("& '") and input_path.endswith("'"):
        path = input_path[3:-1]
    else:
        path = input_path

    # 接下来，去除两端的空格和引号
    return path.strip().strip("'\"")

# 在程序开始时调用清屏
clear_screen()

# 工具介绍和免责声明
print("欢迎使用Excel图片导出工具，请确保所有要导出的图片均为浮动格式，否则无法导出。\n")
print("免责声明：图片导出前请注意Excel文件的备份，使用本程序造成的Excel文件损坏等，均由用户自行承担，与作者无关。\n")

# 输入Excel文件路径
print("请提供Excel文件的完整路径，例如：C:\\Users\\YourName\\Documents\\example.xlsx，可以使用 Ctrl+Z 加回车退出输入。")
while True:
    try:
        raw_path = input("请输入Excel文件的路径: ")
        excel_file_path = clean_path(raw_path)
        # 调试用
        # print(excel_file_path)

        # 验证文件是否存在
        if not os.path.exists(excel_file_path):
            print("指定的文件不存在，请检查路径是否正确，重新输入。\n")
        else:
            break
    except EOFError:
        print("\n输入已取消。")
        exit()

# 获取Excel文件所在的目录
base_output_folder = os.path.dirname(excel_file_path)
print(f"\n基础输出文件夹设定为：{base_output_folder}\n")

# 加载工作簿
print("正在加载工作簿...\n")
wb = load_workbook(excel_file_path, data_only=True)

# 选择sheet
print("请选择图片所在sheet：")
print("1. 导出所有sheet图片")
sheet_names = [name for name in wb.sheetnames if name != "WpsReserved_CellImgList"]  # WPS软件特有的sheet，无实际作用
for idx, name in enumerate(sheet_names, start=2):
    print(f"{idx}. sheet({name})")
selected_sheet_option = input("请输入选项（默认为1）：") or '1'

print("\n")

# 选择图片命名方式
naming_option = input("请选择图片命名方式：\n1. 当前单元格位置\n2. 当前单元格内容\n3. 前一个单元格内容\n请输入选项（默认为1）: ") or '1'

print("\n")

# 选择图片导出格式
export_format = input("请选择图片导出格式：\n1. PNG\n2. JPG\n请输入选项（默认为1）: ") or '1'

print("\n")

# 根据选择的sheet进行导出
sheets_to_export = sheet_names if selected_sheet_option == '1' else [sheet_names[int(selected_sheet_option) - 2]]

# 遍历选定的工作表中的所有图像
print("开始处理图像...\n")
for sheet in sheets_to_export:
    print(f"正在处理工作表：{sheet}\n")

    # 为每个工作表创建一个同名文件夹
    sheet_output_folder = os.path.join(base_output_folder, sheet)
    if not os.path.exists(sheet_output_folder):
        os.makedirs(sheet_output_folder)
        print(f"创建工作表文件夹：{sheet_output_folder}\n")

    ws = wb[sheet]
    for drawing in ws._images:
        if isinstance(drawing, Image):
            # 获取图像所在的单元格
            anchor = drawing.anchor._from
            col_letter = get_column_letter(anchor.col + 1)  # 列增加1
            row_number = anchor.row + 1  # 行增加1
            cell_ref = f"{col_letter}{row_number}"

            # 根据选项确定文件名
            if naming_option == '1':
                file_name = cell_ref
            elif naming_option == '2':
                cell_value = ws[cell_ref].value
                file_name = str(cell_value) if cell_value is not None else cell_ref
            elif naming_option == '3':
                prev_col_letter = get_column_letter(column_index_from_string(col_letter) - 1)
                prev_cell_ref = f"{prev_col_letter}{row_number}"
                prev_cell_value = ws[prev_cell_ref].value
                file_name = str(prev_cell_value) if prev_cell_value is not None else prev_cell_ref
            print(f"发现图像位于：{cell_ref}, 准备处理...")
            
            # 注意：此处的drawing.ref可能无法正确获取嵌入图片的路径
            image_path = drawing.ref
            # 此处假设image_path是正确的图像路径
            try:
                image = PILImage.open(image_path)
                # 根据导出格式选择
                if export_format == '1':
                    ext = '.png'
                    image_format = 'PNG'
                else:
                    ext = '.jpg'
                    image_format = 'JPEG'

                # 生成唯一的保存路径
                save_path = create_unique_filename(os.path.join(sheet_output_folder, file_name), ext)
                image = image.convert('RGB') if ext != '.png' else image
                image.save(save_path, format=image_format)
                print(f"图像保存至：{save_path}\n")
            except IOError:
                print(f"无法打开或保存图像: {image_path}")

print(f"图像导出完成。输出目录：{base_output_folder}\n")
input("按Enter键退出...")
